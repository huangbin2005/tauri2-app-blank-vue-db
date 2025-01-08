# coding=utf-8
__author__ = 'laohuang'

# 如何处理相关
import os
from openpyxl import load_workbook
import re
from datetime import datetime

# 遍历目录
def list_files_and_dirs(directory):
    for name in os.listdir(directory):
        path = os.path.join(directory, name)
        if os.path.isdir(path):
            print(f"Directory: {path}")
        elif os.path.isfile(path):
            # print(f"File: {path}")
            readfile(path)
# 读当月整体考勤
def readfile(file_path):
    workbook = load_workbook(filename=file_path)
    # 文件名提取日期
    filename = file_path
    # 找到下划线之间的部分，并分割
    parts = filename.split('_')
    # 找到包含日期的部分，并再次分割
    date_part = parts[2].split('.')[0]  # 去掉文件扩展名
    start_date, end_date = date_part.split('-')
    # 获取当月开始日期和截至日期
    # print("开始日期:", start_date)
    # print("结束日期:", end_date)
    # 选择工作表
    sheet = workbook['打卡时间']
    # 读取单元格数据
    # cell_value = sheet['A1'].value
    # print(cell_value)
    # 获得当月周末情况
    # 获取第4行数据（openpyxl 的行索引是从1开始的，但我们可以使用 iter_rows 方法并指定 min_row 和 max_row）
    # fourth_row = [cell[0] for cell in sheet.iter_cols(min_col=1, max_col=sheet.max_column, min_row=4, max_row=4, values_only=True)]
    fourth_row = []
    for col in sheet.iter_cols(min_col=1, max_col=sheet.max_column, min_row=4, max_row=4, values_only=True):
        fourth_row.append(col[0])  # 由于 iter_cols 返回的是列的元组，我们取第一个元素（也是唯一一个）
    # 显示第4行数据
    # print(fourth_row)
    weekend_row=[]
    for i in range(6,len(fourth_row)):
        if not (fourth_row[i]).isdigit():
            # print(f"{fourth_row[i]} 不是数字")
            weekend_row.append(i)
    # print(weekend_row)
    # # 遍历所有行和列
    # for row in sheet.iter_rows(values_only=True):
    #     print(row)
    # 从 A5 开始读取（A 列是 1，行是从 5 开始的）
    # 计算起始日期

    start_row = 5  # 行号从 1 开始，但 A5 是第 5 行
    start_col = 1  # A 列是 1,姓名的取值内容
    start_col_01 = 7 # G列是7，第一个考勤的日期

    # 循环读取每一行
    for row in sheet.iter_rows(min_row=start_row, min_col=start_col, values_only=True):
        if row[0]!='ccc' :
            # print(row)  # row 是一个包含该行所有单元格值的元组
            for i in range(start_col_01,len(row)):
                if(row[i] != ""):
                    try :
                        if (i) in weekend_row:
                            print(start_date[:-2]+str(i-start_col_01+2).zfill(2)+"\t周末\t40\t"+row[0]+"\t周末加班\t"+row[0]+"")
                        else:
                            cells = re.sub(r'[^0-9:\n]', '', row[i].replace(" ","")).split("\n")
                            time_format = "%H:%M"
                            parsed_datetime_start = datetime.strptime(cells[0], time_format)
                            parsed_datetime_end = datetime.strptime(cells[len(cells)-1], time_format)
                            time_difference = parsed_datetime_end-parsed_datetime_start
                            total_hours_difference = time_difference.days * 24 + time_difference.seconds // 3600
                            # print(total_hours_difference)
                            if (total_hours_difference > 11 ) or (total_hours_difference<0):
                                # print(cells)
                                print(start_date[:-2]+str(i-start_col_01+2).zfill(2)+"\t周内\t20\t"+row[0]+"\t周内加班\t"+row[0]+"")
                                # print(len(cells))
                        # print(row[i])
                    except Exception as e:
                        print("")

directory_path = "/mnt/d/users/Administrator/Desktop/files/202412/2024年打卡/2024年打卡/"
list_files_and_dirs(directory_path)



